"""Import vendor production-tracking spreadsheets into `factory_production_status`.

WHAT THIS IS FOR: if you manufacture physical goods, your suppliers likely
send you a periodic (often weekly) spreadsheet tracking each style through
production — cut date, sample approval, ex-factory date, shipping status, and
so on. Different vendors virtually never agree on a column layout: different
column names, different column ORDER, sometimes a different layout per SHEET
within one vendor's own file (e.g. one tab per product category). This
connector normalizes an arbitrary mix of such spreadsheets into ONE table via
KEYWORD-BASED header matching rather than hardcoded per-file column indices,
because header wording drifts week to week (line breaks, trailing translated
text, minor rewording) even from the same vendor.

This is a MANUAL-DROP importer, the same shape as `voc_import.py`/
`pacvue_import.py` elsewhere in this repo: someone downloads/receives the
files and drops them into a folder; you run this script by hand (or on your
own schedule) to load them. There's no API here — production-tracking sheets
essentially never have one.

HOW HEADER MATCHING WORKS: `FIELD_RULES` is an ORDERED, most-specific-first
list of (canonical_field, [regex, ...]) pairs. Each column header is
normalized (lowercased, whitespace collapsed, non-Latin trailing translation
text stripped) and matched against these patterns in order; the first match
wins. Order matters — e.g. "actual cut date" must be checked before the
bare "cut date" pattern, or the more specific field never gets a chance to
match. Any column that doesn't match anything is preserved verbatim in a
JSON `extra_json` sidecar column rather than silently dropped — this is a
first pass at unifying a genuinely messy set of vendor templates, and you
should expect to grow `FIELD_RULES` as you encounter more vendors/columns.

OPTIONAL JOIN-BACK TO A PRODUCT MASTER: if your warehouse has some other
table mapping a style/model number to your own internal product identity
(a NetSuite export, a hand-maintained product sheet, whatever you call your
own "source of truth" product table), `resolve_matches()` can join this
sheet's `style_no` column against it — see `PRODUCT_MASTER_TABLE` /
`PRODUCT_MASTER_STYLE_COLUMN` below. This repo ships no such table by
default, so the join is skipped cleanly (empty match dict, no error) unless
you point it at a real table with a compatible shape. A style still in
production and not yet in your product master is expected NOT to match yet —
rows are kept with their match columns left NULL rather than dropped, since a
later re-run naturally re-links them once the style is added to your master
table.

Usage:
    factory_status_import.py <folder>            # one period's folder, all vendor files
    factory_status_import.py <folder> --dry-run   # parse + report, no writes
"""

import argparse
import json
import os
import re
import sqlite3
import sys
from datetime import datetime, date

import openpyxl

try:
    import xlrd
except ImportError:
    xlrd = None

DB_PATH = os.environ.get("WAREHOUSE_DB", os.path.join(os.path.dirname(__file__), "warehouse.db"))

# Optional join-back target — see "OPTIONAL JOIN-BACK" in the module
# docstring. Point these at your own product-master table/columns if you
# have one; leave PRODUCT_MASTER_TABLE as-is (or set it to "") to disable
# the join entirely. resolve_matches() checks the table actually exists
# before querying it, so an unset/wrong table degrades to "no matches"
# rather than erroring.
PRODUCT_MASTER_TABLE = "product_master"
PRODUCT_MASTER_STYLE_COLUMN = "style_no"       # column holding the style/model number
PRODUCT_MASTER_SNAPSHOT_COLUMN = "snapshot_date"  # set to "" if the table isn't snapshot-keyed
PRODUCT_MASTER_SKU_COLUMN = "sku"
PRODUCT_MASTER_GROUP_COLUMN = "product_group"  # e.g. a color-family/style-group identifier
PRODUCT_MASTER_ID_COLUMN = "product_id"

# ---------------------------------------------------------------------------
# Canonical schema + header -> canonical field matching
# ---------------------------------------------------------------------------

# Ordered (most-specific-first) keyword rules. A header is normalized (lower,
# collapse whitespace, strip trailing non-Latin text) then matched against
# these in order; first hit wins. Order matters: "actual cut date" must be
# checked before the bare "cut date" pattern, etc. Extend this list for your
# own vendors' column vocabulary — it's expected to grow.
FIELD_RULES = [
    ("style_no", [r"\bstyle\s*no\b", r"\bstyle#", r"\bstyle\s*number\b", r"^model\s*no\b", r"^mpn$"]),
    ("description", [r"\bdescription\b", r"\bstyle\s*name\b", r"\bproduct description\b"]),
    ("brand", [r"^brand$"]),
    ("collection_month", [r"collection month", r"month\s*&?\s*year"]),
    ("designer", [r"^designer$"]),
    ("buyer", [r"^buyer$"]),
    ("colour", [r"^colou?r$", r"material/?color approve"]),
    ("sizes", [r"\bsizes?\b.*regular", r"^size$"]),
    ("supplier", [r"^supplier$"]),
    ("category", [r"^category$"]),
    ("subcategory", [r"^subcat$"]),
    ("actual_ex_factory_date", [r"actual.*ex-?factory"]),
    ("original_ex_factory_date", [r"original.*ex-?factory", r"planned.*vessel.*book"]),
    ("actual_cut_date", [r"actual.*cut date"]),
    ("planned_cut_date", [r"planned.*cut date"]),
    ("actual_packing_date", [r"actual packing date"]),
    ("planned_packing_date", [r"planned packing date"]),
    ("actual_vessel_book_date", [r"actual.*vessel book"]),
    ("planned_vessel_book_date", [r"planned.*vessel book"]),
    ("planned_wh_date", [r"planned wh date", r"planned\s*indc date"]),
    ("cancel_date_on_po", [r"cancel date"]),
    ("po_number", [r"^po#?\s*$", r"\bpo number\b", r"^po\s*#"]),
    ("po_issue_date", [r"po issue date", r"received po date"]),
    ("tech_pack_received_date", [r"tech pack.*received", r"tech pack.*rec"]),
    ("original_sample_received_date", [r"original sample received", r"orig sample.*rec"]),
    ("artwork_received_date", [r"artwork received"]),
    ("sample_send_date", [r"^sample send date", r"tp/orig sample send date"]),
    ("fit_comments_rcvd_date", [r"fit comments"]),
    ("revised_fit_sample_send_date", [r"revised fit sample"]),
    ("final_fit_approved", [r"final fit approved"]),
    ("color_approved", [r"color approved"]),
    ("print_embellishment_approved", [r"print/?embellishment"]),
    ("branding_packaging", [r"branding/?packaging"]),
    ("top_send_date_awb", [r"top send date", r"top to \w+"]),
    ("length_in", [r"length\s*\(in"]),
    ("width_in", [r"width\s*\(in", r"cuttable width"]),
    ("height_in", [r"height\s*\(in"]),
    ("weight_lb", [r"weight\s*\(lb"]),
    ("coo", [r"^coo$"]),
    ("hts", [r"^hts$"]),
    ("fabric_type", [r"fabric.*type"]),
    ("fabric_content", [r"fabric.*content"]),
    ("fabric_cost_usd", [r"fabric cost"]),
    ("yy_yards", [r"^yy\b", r"yards?\)?$"]),
    ("units", [r"^units?$", r"^unite$"]),
    ("actual_shipped", [r"actual shipped"]),
    ("price_usd", [r"\bprice\b", r"ddp cost", r"fob price"]),
    ("total_amount", [r"total amount"]),
    ("xf_date", [r"^xf date$"]),
    ("vendor", [r"^vendor$"]),
    ("bulk_factory_name", [r"bulk factory name"]),
    ("fob_port", [r"fob port"]),
    ("ship_mode", [r"ship mode", r"ship.*carrier"]),
    ("shipping_status", [r"shipping status"]),
    ("drop_flag", [r"^drop\b"]),
    ("vendor_comments", [r"vendor comments"]),
    ("internal_comments", [r"\bcomments\b"]),
    ("sku_number", [r"^sku#?$"]),
    ("photo", [r"^photo$", r"^image$"]),
]

CANONICAL_FIELDS = [f for f, _ in FIELD_RULES]

DATE_FIELDS = {
    "tech_pack_received_date", "original_sample_received_date", "artwork_received_date",
    "sample_send_date", "fit_comments_rcvd_date", "revised_fit_sample_send_date",
    "final_fit_approved", "color_approved", "print_embellishment_approved",
    "branding_packaging", "top_send_date_awb", "cancel_date_on_po", "planned_wh_date",
    "planned_cut_date", "actual_cut_date", "planned_packing_date", "actual_packing_date",
    "original_ex_factory_date", "actual_ex_factory_date", "planned_vessel_book_date",
    "actual_vessel_book_date", "po_issue_date", "xf_date", "collection_month",
}


def normalize_header(h):
    if h is None:
        return ""
    s = str(h)
    # Strip trailing non-Latin translation text some vendor templates append
    # to an otherwise-English header (e.g. a Chinese or Korean translation in
    # the same cell) -- generalize this range if your own vendors use a
    # different script.
    s = re.sub(r"[一-鿿가-힣]+", "", s)
    s = s.replace("\n", " ")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def match_field(norm_header):
    for field, patterns in FIELD_RULES:
        for pat in patterns:
            if re.search(pat, norm_header):
                return field
    return None


def build_column_map(headers):
    """Map column index -> canonical field name (first match wins per field)."""
    col_map = {}
    used_fields = set()
    for idx, h in enumerate(headers):
        norm = normalize_header(h)
        if not norm:
            continue
        field = match_field(norm)
        if field and field not in used_fields:
            col_map[idx] = field
            used_fields.add(field)
    return col_map


def excel_serial_to_iso(val):
    """Convert an Excel serial date number to ISO date string, else pass through."""
    if isinstance(val, (int, float)):
        try:
            base = date(1899, 12, 30)
            d = base + __import__("datetime").timedelta(days=float(val))
            if 1990 <= d.year <= 2035:
                return d.isoformat()
        except Exception:
            return val
    if isinstance(val, (datetime, date)):
        return val.isoformat() if isinstance(val, datetime) else val.isoformat()
    return val


def clean_value(field, val):
    if val in (None, "", "#VALUE!", "#N/A", "#REF!"):
        return None
    if field in DATE_FIELDS:
        val = excel_serial_to_iso(val)
    # SQLite's sqlite3 driver can't bind datetime.time (or any other
    # non-primitive) at all -- stringify defensively regardless of field,
    # since a mis-detected header can route a time-formatted cell anywhere.
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    if isinstance(val, __import__("datetime").time):
        return val.isoformat()
    if isinstance(val, str):
        val = val.strip()
        if val == "":
            return None
    return val


def find_header_row(rows, max_scan=6):
    """Pick the row (within the first max_scan) with the most string-like header cells."""
    best_row, best_score = 0, -1
    for i, row in enumerate(rows[:max_scan]):
        score = sum(1 for c in row if isinstance(c, str) and len(c.strip()) > 2)
        if score > best_score:
            best_score, best_row = score, i
    return best_row


# ---------------------------------------------------------------------------
# File readers
# ---------------------------------------------------------------------------

def iter_sheets_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        if not hasattr(ws, "iter_rows"):
            continue  # Chartsheet or other non-worksheet object, not product data
        rows = list(ws.iter_rows(values_only=True))
        yield sn, rows


def iter_sheets_xls(path):
    if xlrd is None:
        raise RuntimeError("xlrd not installed; cannot read legacy .xls files")
    wb = xlrd.open_workbook(path)
    for sn in wb.sheet_names():
        sh = wb.sheet_by_name(sn)
        rows = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
        yield sn, rows


def parse_file(path, vendor_hint, snapshot_date):
    """Yield normalized record dicts for every product row in every usable sheet."""
    fn = os.path.basename(path)
    ext = fn.lower().rsplit(".", 1)[-1]
    try:
        sheets = iter_sheets_xlsx(path) if ext == "xlsx" else iter_sheets_xls(path)
    except Exception as e:
        print(f"  ! failed to open {fn}: {e}", file=sys.stderr)
        return

    for sheet_name, rows in sheets:
        if len(rows) < 2:
            continue
        hdr_idx = find_header_row(rows)
        headers = rows[hdr_idx]
        col_map = build_column_map(headers)
        mapped_fields = set(col_map.values())
        # need at least a style/description anchor to trust this sheet as product data
        if "style_no" not in mapped_fields and "description" not in mapped_fields:
            continue

        for row in rows[hdr_idx + 1:]:
            if not any(c not in (None, "") for c in row):
                continue
            rec = {f: None for f in CANONICAL_FIELDS}
            extra = {}
            for idx, val in enumerate(row):
                if val in (None, ""):
                    continue
                field = col_map.get(idx)
                if field:
                    cleaned = clean_value(field, val)
                    if cleaned is not None and rec.get(field) is None:
                        rec[field] = cleaned
                else:
                    hdr = normalize_header(headers[idx]) if idx < len(headers) else f"col_{idx}"
                    if hdr:
                        extra[hdr] = val

            style_no = rec.get("style_no")
            if style_no is not None:
                style_no = re.sub(r"\s*-\s*[A-Za-z ]+$", "", str(style_no)).strip()  # strip trailing " - RED"
                style_no = style_no.strip() or None
            if not style_no and not rec.get("description"):
                continue

            rec["style_no"] = style_no
            rec["source_file"] = fn
            rec["source_vendor"] = vendor_hint
            rec["source_sheet"] = sheet_name
            rec["snapshot_date"] = snapshot_date
            rec["extra_json"] = json.dumps(extra, default=str) if extra else None
            yield rec


# EDIT THIS FOR YOUR OWN VENDORS. Each entry is (filename-matching regex,
# display name); the first pattern that matches the (lowercased) filename
# wins, and an unmatched filename falls back to its own stem as the vendor
# name. The example patterns below are placeholders -- replace them with
# your own suppliers'.
VENDOR_FROM_FILENAME = [
    (r"vendor[_ -]?a", "Vendor A"),
    (r"apparel.*vendor[_ -]?b|vendor[_ -]?b.*apparel", "Vendor B - Apparel"),
    (r"footwear.*vendor[_ -]?b|vendor[_ -]?b.*footwear", "Vendor B - Footwear"),
    (r"vendor[_ -]?b", "Vendor B - Other"),
    (r"vendor[_ -]?c", "Vendor C"),
]


def vendor_from_filename(fn):
    low = fn.lower()
    for pat, name in VENDOR_FROM_FILENAME:
        if re.search(pat, low):
            return name
    return os.path.splitext(fn)[0]


DDL = """
CREATE TABLE IF NOT EXISTS factory_production_status (
    snapshot_date TEXT NOT NULL,
    source_vendor TEXT NOT NULL,
    source_file TEXT NOT NULL,
    source_sheet TEXT NOT NULL,
    row_num INTEGER NOT NULL,
    style_no TEXT,
    description TEXT,
    brand TEXT,
    collection_month TEXT,
    designer TEXT,
    buyer TEXT,
    colour TEXT,
    sizes TEXT,
    supplier TEXT,
    category TEXT,
    subcategory TEXT,
    tech_pack_received_date TEXT,
    original_sample_received_date TEXT,
    artwork_received_date TEXT,
    sample_send_date TEXT,
    fit_comments_rcvd_date TEXT,
    revised_fit_sample_send_date TEXT,
    final_fit_approved TEXT,
    color_approved TEXT,
    print_embellishment_approved TEXT,
    branding_packaging TEXT,
    top_send_date_awb TEXT,
    length_in REAL,
    width_in REAL,
    height_in REAL,
    weight_lb REAL,
    coo TEXT,
    hts TEXT,
    fabric_type TEXT,
    fabric_content TEXT,
    fabric_cost_usd REAL,
    yy_yards REAL,
    units REAL,
    actual_shipped REAL,
    price_usd TEXT,
    total_amount REAL,
    po_number TEXT,
    po_issue_date TEXT,
    cancel_date_on_po TEXT,
    planned_wh_date TEXT,
    planned_cut_date TEXT,
    actual_cut_date TEXT,
    planned_packing_date TEXT,
    actual_packing_date TEXT,
    original_ex_factory_date TEXT,
    actual_ex_factory_date TEXT,
    xf_date TEXT,
    planned_vessel_book_date TEXT,
    actual_vessel_book_date TEXT,
    vendor TEXT,
    bulk_factory_name TEXT,
    fob_port TEXT,
    ship_mode TEXT,
    shipping_status TEXT,
    drop_flag TEXT,
    vendor_comments TEXT,
    internal_comments TEXT,
    sku_number TEXT,
    photo TEXT,
    extra_json TEXT,
    -- OPTIONAL join-back to your own product-master table (see the module
    -- docstring / PRODUCT_MASTER_* constants). Left NULL when unmatched or
    -- when no product-master table is configured.
    matched_sku_key TEXT,
    matched_product_group TEXT,
    matched_product_id TEXT,
    match_method TEXT,
    PRIMARY KEY (snapshot_date, source_file, source_sheet, row_num)
)
"""

STORE_FIELDS = [f for f in CANONICAL_FIELDS if f not in ("style_no",)]


def _table_exists(conn, table: str) -> bool:
    return conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (table,)
    ).fetchone() is not None


def resolve_matches(conn, records):
    """Look up style_no against an optional product-master table (see
    PRODUCT_MASTER_* constants), one query for all styles. Returns {} with no
    error if PRODUCT_MASTER_TABLE is unset or the table doesn't exist in this
    warehouse -- the join-back is optional, and this repo ships no such table
    by default."""
    if not PRODUCT_MASTER_TABLE or not _table_exists(conn, PRODUCT_MASTER_TABLE):
        return {}
    styles = sorted(set(r["style_no"] for r in records if r.get("style_no")))
    if not styles:
        return {}
    match = {}
    CHUNK = 400
    cur = conn.cursor()
    snapshot_clause = (
        f"AND {PRODUCT_MASTER_SNAPSHOT_COLUMN} = "
        f"(SELECT MAX({PRODUCT_MASTER_SNAPSHOT_COLUMN}) FROM {PRODUCT_MASTER_TABLE})"
        if PRODUCT_MASTER_SNAPSHOT_COLUMN else ""
    )
    for i in range(0, len(styles), CHUNK):
        chunk = styles[i:i + CHUNK]
        qmarks = ",".join("?" for _ in chunk)
        try:
            cur.execute(f"""
                SELECT {PRODUCT_MASTER_STYLE_COLUMN},
                       MIN({PRODUCT_MASTER_SKU_COLUMN}) AS sku_key,
                       MIN({PRODUCT_MASTER_GROUP_COLUMN}) AS product_group,
                       MIN({PRODUCT_MASTER_ID_COLUMN}) AS product_id
                FROM {PRODUCT_MASTER_TABLE}
                WHERE {PRODUCT_MASTER_STYLE_COLUMN} IN ({qmarks})
                {snapshot_clause}
                GROUP BY {PRODUCT_MASTER_STYLE_COLUMN}
            """, chunk)
        except sqlite3.OperationalError:
            # A configured product-master table with a different column shape
            # than expected -- degrade to "no matches" rather than crashing
            # an otherwise-successful import.
            return {}
        for style_no, sku_key, product_group, product_id in cur.fetchall():
            match[style_no] = (sku_key, product_group, product_id)
    return match


def snapshot_date_from_folder(folder):
    folder_name = os.path.basename(folder.rstrip("\\/"))
    m = re.match(r"(\d{2})\.(\d{2})\.(\d{4})", folder_name)
    if m:
        mm, dd, yyyy = m.groups()
        return f"{yyyy}-{mm}-{dd}"
    return None


def process_folder(folder, conn, snapshot_date=None, dry_run=False, verbose=True):
    """Parse every vendor file in one dated period folder and load it into
    factory_production_status on the given (already-open) connection. Returns
    a stats dict. Never raises on a single bad file -- records the error and
    keeps going.
    """
    snapshot_date = snapshot_date or snapshot_date_from_folder(folder) or datetime.now().date().isoformat()
    # Some periods drop files flat in the dated folder; others nest them
    # under category subfolders (e.g. apparel/footwear/accessories). Walk
    # recursively and use the path relative to `folder` as the file identity,
    # since several subfolders can reuse the same filename for the same
    # vendor's submission to a different category.
    rel_paths = []
    for root, _dirs, filenames in os.walk(folder):
        for fn in filenames:
            if fn.lower().endswith((".xlsx", ".xls")):
                rel_paths.append(os.path.relpath(os.path.join(root, fn), folder))
    rel_paths.sort()
    if verbose:
        print(f"Folder: {folder}\nSnapshot date: {snapshot_date}\nFiles: {len(rel_paths)}")

    all_records = []
    file_errors = {}
    for rel in rel_paths:
        path = os.path.join(folder, rel)
        fn = os.path.basename(rel)
        subfolder = os.path.dirname(rel)  # '' if flat, else 'apparel' / 'footwear' / ...
        vendor = vendor_from_filename(fn)
        try:
            recs = list(parse_file(path, vendor, snapshot_date))
        except Exception as e:
            file_errors[rel] = str(e)
            if verbose:
                print(f"  ! {rel} FAILED: {e}", file=sys.stderr)
            continue
        for r in recs:
            r["source_file"] = rel  # keep the subfolder in the identity to avoid PK collisions
            if subfolder and not r.get("category"):
                r["category"] = subfolder
        all_records.extend(recs)
        if verbose:
            print(f"  {rel}  -> {len(recs)} rows  (vendor={vendor})")

    matches = resolve_matches(conn, all_records)
    n_matched = 0
    for r in all_records:
        m = matches.get(r.get("style_no"))
        if m:
            r["matched_sku_key"], r["matched_product_group"], r["matched_product_id"] = m
            r["match_method"] = "style_exact"
            n_matched += 1
        else:
            r["matched_sku_key"] = r["matched_product_group"] = r["matched_product_id"] = None
            r["match_method"] = None

    stats = {
        "snapshot_date": snapshot_date,
        "folder": folder,
        "n_files": len(rel_paths),
        "n_file_errors": len(file_errors),
        "file_errors": file_errors,
        "n_rows": len(all_records),
        "n_distinct_styles": len(set(r["style_no"] for r in all_records if r.get("style_no"))),
        "n_matched": n_matched,
    }

    if dry_run or not all_records:
        return stats

    cur = conn.cursor()
    cur.execute(DDL)
    cur.execute("DELETE FROM factory_production_status WHERE snapshot_date = ?", (snapshot_date,))

    cols = ["snapshot_date", "source_vendor", "source_file", "source_sheet", "row_num", "style_no"] + \
           STORE_FIELDS + ["matched_sku_key", "matched_product_group", "matched_product_id", "match_method"]
    placeholders = ",".join("?" for _ in cols)
    sql = f"INSERT OR REPLACE INTO factory_production_status ({','.join(cols)}) VALUES ({placeholders})"

    row_counters = {}
    insert_rows = []
    for r in all_records:
        key = (r["source_file"], r["source_sheet"])
        row_counters[key] = row_counters.get(key, 0) + 1
        vals = [r["snapshot_date"], r["source_vendor"], r["source_file"], r["source_sheet"],
                 row_counters[key], r.get("style_no")]
        vals += [r.get(f) for f in STORE_FIELDS]
        vals += [r["matched_sku_key"], r["matched_product_group"], r["matched_product_id"], r["match_method"]]
        insert_rows.append(vals)

    cur.executemany(sql, insert_rows)
    conn.commit()
    if verbose:
        print(f"\nWrote {len(insert_rows)} rows to factory_production_status (snapshot_date={snapshot_date})")
    return stats


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("folder", help="a dated period folder, e.g. .../08.24.2026")
    ap.add_argument("--snapshot-date", help="override snapshot date (default: parsed from folder name)")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    conn = sqlite3.connect(DB_PATH)
    stats = process_folder(args.folder, conn, snapshot_date=args.snapshot_date, dry_run=args.dry_run)
    print(f"\nTotal rows parsed: {stats['n_rows']}")
    print(f"Distinct style_no: {stats['n_distinct_styles']}")
    if stats["n_rows"]:
        print(f"Matched to product master: {stats['n_matched']} / {stats['n_rows']} "
              f"({stats['n_matched'] / stats['n_rows'] * 100:.1f}%)")
    if stats["file_errors"]:
        print(f"File errors: {stats['file_errors']}")
    conn.close()


if __name__ == "__main__":
    main()
